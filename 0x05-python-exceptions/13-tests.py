#python exception handling

#4-square
"""i = 0
while(i <= 10):
    print(i * "*")
    i += 1

import time
for seconds in range(10,0,-1):
    print("{}".format(seconds))
    time.sleep(1)
print("Happy new year")"""

"""import time
for i in range(20):
    print(i, end="\n")
    time.sleep(1)
print("Happy new Year")"""

#fizz buzz
"""def fizz_buzz(x):
    if ((x % 3 == 0) and (x % 5 == 0)):
        return "FizzBuzz"
    if x % 3 == 0:
        return "Fizz"
    if x % 5 == 0:
        return "Buzz"
    return x
print(fizz_buzz(4))
print(fizz_buzz(3))
print(fizz_buzz(5))
print(fizz_buzz(15))"""

#2D lists
"""names = ["Emmanuel", "Wambetsa", "Shitseswa"]
school = ["Primary", "Secondary", "University"]
my_details = [names, school]
for i in my_details:
    for j in i:
        print(j)
    print()"""

#tuples
"""names = ("Emmanuel", "Wambetsa", "Shitseswa")
school = ("Primary", "Secondary", "University")
print(names.count("Emmanuel"))
print(names.index("Wambetsa"))
if "Emmanuel" in names:
    ind = names.count
    print(f"Emmanuel appears in names")"""

#set
"""utensils = {"fork", "spoon", "knives"}
dishes = {"plate", "cup", "play", "knives"}
for x in utensils:
    print(x)
print(utensils.union(dishes))
print(utensils.difference(dishes))"""

#dictionaries
"""capitals = {"USA":"DC", "Kenya":"Nairobi", "Rusia": "Moscow", "China":"Beijing"}
print(capitals.get("Uganda"))
print(capitals.keys())
print(capitals.values())
print(capitals.items())
capitals.update({"SA":"Pretoria"})
for k,v in capitals.items():
    print(k,v)
capitals.clear()"""

#Index operator
"""name = "Emmanuel wambetsa"
if(name[0].islower()):
    name = name.capitalize()
print(name)
f_name = name[0:9].upper()
l_name = name[9:].capitalize()
rev_name = name[::-1]
print(f_name)
print(l_name)
print(rev_name)"""

#functions - executed when called. params arguments - values of params
"""def hello(fn, ln):
    print("Hello: {} {}".format(fn, ln))
hello("Dev", "Wambetsa")"""

#return stt
"""def mult(a, b):
    return (a * b)
print(mult(3,5))"""

#keyword arguments

#nested function calls
"""print(round(abs(float(input("Enter number: ")))))"""

#variable scope - local vs global variables - LEGB

#*args parameter - tuples params
"""def add(*args):
    sum = 0
    for i in args:
        sum += i
    return sum
print(add(3,4,5,6,7,8,9))"""

#changing vals in tupple *arg by casting to list b4 changing
"""def add(*args):
    sum = 0
    args = list(args)
    args[0] = 100
    for i in args:
        sum += i
    return sum
print(add(3,4,5,6,7,8,9))"""

#**kwargs = changes args to dictionary
"""def hello(**kwargs):
    print("Hello", end=" ")
    for k,v in kwargs.items():
        print(v, end=" ")

hello(title="Mr", fn="Emmanuel", mn="Wambetsa", ln="Shitseswa")"""

#string format
"""animal = "cow"
item = "moon"
print("The {} jumped over the {}".format(animal, item))
print("The {1} jumped over the {0}".format(animal, item))#positional argument
print("The {anim} jumped over the {itm}".format(anim="pig", itm="sun"))#keyword argument
print(f"The {animal} jumped over the {item}")

text = "The {} score a goal on the {} thrice"
print(text.format(animal, item))


name = "Emmanuel"
print("Hello, my name is {:18}. Nice to meet you".format(name))
print("Hello, my name is {:<18}. Nice to meet you".format(name))
print("Hello, my name is {:>18}. Nice to meet you".format(name))
print("Hello, my name is {:^18}. Nice to meet you".format(name))
print("Hello, my name is {0:<18}. Nice to meet you".format(name))#positional argument"""

#exception handling ==== try, except1, except2, exceptn, else:print result, finally



"""class Solution:
    def twoSum(self, nums: List[int], target: int) -> List[int]:
        for i in range(len(nums)):

            j = i + 1
            for j in range(len(nums)):
                if nums[i] + nums[j] == target:
                    return(i, j)"""
        
###openpyxl
"""import openpyxl
inventory_file_name = openpyxl.load_workbook("inventories.xlsx")
prod_list = inventory_file_name["Sheet1"]
products_per_supplier = {}

for product_row in range(2, prod_list.max_row + 1):
    supplier_name = prod_list.cell(product_row, 4)"""

a = 1
b = a
a = 2
print(b)
l = [1,2,3]
m = l
print(m)
l[2] = 'x'
print(m)

"""Aliasing"""
names = ["Emmanuel", "Wambetsa"]
my_names = names
print(my_names)
names.append("Shitseswa")
print(names)
print(my_names)
"""cloning lists"""
a = [1,2,3,4,5,6,7]
b = a[:]
print(b)
b[0] = 10
print(b)
"""rev"""
print(a[::-1])

a = 89
b = a + 1
print(a is b)

#What does this script print?

def increment(n):
    n.append(4)

l = [1, 2, 3]
increment(l)
print(l)
#What does this script print?

a = (1)
b = (1)
print(a is b)

import random
choices = ["Chelsea", "Man United", "Man City", "Arsenal"]
res = random.choice(choices)
user_choice = None

while user_choice not in choices:
    user_choice = input("Which team is lucky? : ").lower()

#if user_choice == res:
print("Won, Lucky team is: {}".format(user_choice))